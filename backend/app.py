"""
app.py
Backend Flask do Sistema de Gestão de Eventos e Check-in de Convidados.

Expõe uma API REST consumida pelo front-end web (servido pelo próprio
Flask), utilizada tanto pelo Painel Administrativo quanto pelo Painel
do Porteiro. Roda em 0.0.0.0 para que múltiplos tablets na mesma rede
local possam se conectar ao mesmo backend simultaneamente.
"""

import os
import re
import io
import base64
import hashlib
import traceback
from datetime import timedelta
from functools import wraps

from flask import (
    Flask, request, jsonify, render_template, send_file, send_from_directory, abort,
    session, redirect, url_for,
)
from flask_cors import CORS
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill
from PIL import Image, UnidentifiedImageError

import config
import database as db
import email_service
import event_status
import sms_service
import whatsapp_service
from security import login_attempt_tracker
from text_utils import slugify_filename, normalize_msisdn
from ssl_utils import ensure_self_signed_cert
from qrcode_utils import generate_guest_qrcode, extract_guest_id_from_payload, QRCodeGenerationError
from pdf_generator import generate_invites_pdf, generate_contingency_pdf, PDFGenerationError
from reports import generate_attendance_report_csv

config.ensure_directories()

app = Flask(
    __name__,
    template_folder=os.path.join(os.path.dirname(config.BASE_DIR), "frontend", "templates"),
    static_folder=os.path.join(os.path.dirname(config.BASE_DIR), "frontend", "static"),
)
app.config["MAX_CONTENT_LENGTH"] = config.MAX_UPLOAD_SIZE_BYTES

# Fase 5: pasta de assets estáticos da marca (logo etc.), servida via rota
# dedicada (não fica dentro de frontend/static para o usuário poder trocar
# o arquivo sem mexer no código do frontend).
ASSETS_DIR = os.path.join(config.BASE_DIR, "assets")
os.makedirs(ASSETS_DIR, exist_ok=True)
GATEFLOW_LOGO_FILENAME = "GateFlow.png"
GATEFLOW_LOGO_PATH = os.path.join(ASSETS_DIR, GATEFLOW_LOGO_FILENAME)


@app.route("/assets/<path:filename>")
def serve_asset(filename):
    """Serve arquivos de marca (logo etc.) de checkin_system/backend/assets/."""
    return send_from_directory(ASSETS_DIR, filename)


@app.route("/sw.js")
def serve_service_worker():
    """
    Serve o Service Worker do check-in offline-first a partir da RAIZ do
    site (não de /static/sw.js) -- necessário para o seu escopo cobrir
    páginas como /checkin/<id>, não só /static/. O cabeçalho
    Service-Worker-Allowed reforça isso mesmo servindo o ficheiro físico
    de dentro da pasta static/.
    """
    response = send_from_directory(app.static_folder, "sw.js")
    response.headers["Service-Worker-Allowed"] = "/"
    response.headers["Cache-Control"] = "no-cache"
    return response


@app.context_processor
def inject_brand_assets():
    """Disponibiliza gateflow_logo_available em TODOS os templates, sem precisar passar em cada render_template."""
    return {"gateflow_logo_available": os.path.exists(GATEFLOW_LOGO_PATH)}


# --- Fase 5: Rate Limiting (Flask-Limiter) ---
# Protege rotas publicas de abuso/DoS: checkout, webhook M-Pesa e login sao
# os alvos mais sensiveis (custam dinheiro/credenciais se abusados). O
# armazenamento padrao e em memoria (adequado para 1 processo); em producao
# com multiplos workers, configurar RATELIMIT_STORAGE_URI para um Redis
# compartilhado (ver .env.example).
try:
    from flask_limiter import Limiter
    from flask_limiter.util import get_remote_address

    limiter = Limiter(
        get_remote_address,
        app=app,
        default_limits=["600 per hour"],
        storage_uri=os.environ.get("RATELIMIT_STORAGE_URI", "memory://"),
    )
except ImportError:  # pragma: no cover - so acontece se Flask-Limiter nao estiver instalado
    app.logger.warning("Flask-Limiter nao instalado -- rate limiting DESATIVADO. Instale via requirements.txt.")

    class _NoOpLimiter:
        def limit(self, *_a, **_k):
            def decorator(f):
                return f
            return decorator

    limiter = _NoOpLimiter()

# Em produção (Vercel) vem da variável de ambiente SECRET_KEY; em
# desenvolvimento local, persistida em disco (data/secret.key) para que
# os logins não sejam invalidados a cada reinício do servidor.
app.secret_key = config.get_or_create_secret_key()
app.permanent_session_lifetime = timedelta(minutes=config.SESSION_LIFETIME_MINUTES)

CORS(app, supports_credentials=True)

# Em serverless (Vercel), este módulo é importado a cada cold start, o
# que pode significar múltiplas instâncias rodando init_db() quase ao
# mesmo tempo contra o Neon. init_db() é idempotente (tudo é
# "IF NOT EXISTS"/checagem de coluna antes de alterar), então isso é
# seguro -- mas uma disputa transitória de lock no Postgres não deve
# derrubar a função inteira. Se isso acontecer, o app.logger registra o
# erro e a aplicação segue no ar (as tabelas já devem existir de um
# cold start anterior ou da migração manual feita no deploy).
try:
    db.init_db()
except Exception:  # noqa: BLE001 - não deixamos um cold start concorrente derrubar o app
    app.logger.exception("Falha ao rodar init_db() neste cold start -- verifique DATABASE_URL/Neon.")


def _seed_default_users():
    """
    Cria os usuários padrão (admin e porteiro) apenas na primeiríssima
    execução, quando a tabela de usuários ainda está vazia. NÃO afeta
    eventos/convidados já existentes. As senhas padrão são exibidas uma
    única vez no console e DEVEM ser trocadas imediatamente em produção.
    """
    if db.count_users() > 0:
        return
    admin_hash = generate_password_hash(config.DEFAULT_ADMIN_PASSWORD)
    porteiro_hash = generate_password_hash(config.DEFAULT_PORTEIRO_PASSWORD)
    db.create_user(config.DEFAULT_ADMIN_USERNAME, admin_hash, "admin")
    db.create_user(config.DEFAULT_PORTEIRO_USERNAME, porteiro_hash, "porteiro")
    print("\n" + "=" * 72)
    print("!! USUÁRIOS PADRÃO CRIADOS — TROQUE AS SENHAS IMEDIATAMENTE !!")
    print(f"   Admin      -> usuário: {config.DEFAULT_ADMIN_USERNAME}    senha: {config.DEFAULT_ADMIN_PASSWORD}")
    print(f"   Porteiro   -> usuário: {config.DEFAULT_PORTEIRO_USERNAME} senha: {config.DEFAULT_PORTEIRO_PASSWORD}")
    print("=" * 72 + "\n")


try:
    _seed_default_users()
except Exception:  # noqa: BLE001 - mesmo raciocínio do try/except em torno de db.init_db() acima
    app.logger.exception("Falha ao semear usuários padrão neste cold start.")


# --------------------------------------------------------------------------
# AUTENTICAÇÃO E CONTROLE DE ACESSO (RBAC)
# --------------------------------------------------------------------------

def get_current_user():
    """
    Retorna o usuario logado (dict) a partir da sessao, ou None.

    Fase 5: tambem aplica sessao unica por utilizador -- se o token
    guardado nesta sessao nao bater com o token ativo do usuario no
    banco (porque ele fez login em outro dispositivo/navegador depois
    desta sessao ter sido criada), a sessao e tratada como invalida e
    encerrada aqui mesmo.
    """
    user_id = session.get("user_id")
    if not user_id:
        return None
    session_token = session.get("session_token")
    if not db.is_session_token_valid(user_id, session_token):
        session.clear()
        return None
    return db.get_user_by_id(user_id)


def login_required_page(view_func):
    """Protege rotas HTML: redireciona para /login se não autenticado."""
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not get_current_user():
            return redirect(url_for("page_login", next=request.path))
        return view_func(*args, **kwargs)
    return wrapper


def _porteiro_can_access_event(user, event):
    """
    Regra de isolamento de porteiros (Item C da Fase 5):
      - admin: acesso a TODOS os eventos.
      - porteiro com organizador_id nulo ("porteiro da empresa" / global,
        criado pelo Super Admin): acesso a TODOS os eventos.
      - porteiro com organizador_id preenchido ("porteiro do promotor",
        criado por um Organizador): acesso SOMENTE aos eventos cujo
        organizador_id bata com o do seu criador.
    """
    if not user or not event:
        return False
    if user["role"] == "admin":
        return True
    if user["role"] == "porteiro":
        if not user.get("organizador_id"):
            return True  # porteiro global (da empresa)
        return event.get("organizador_id") == user["organizador_id"]
    return False


def role_required_page(*allowed_roles):
    """Protege rotas HTML por papel (role). Barra com 403 se o papel não bate."""
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(*args, **kwargs):
            user = get_current_user()
            if not user:
                return redirect(url_for("page_login", next=request.path))
            if user["role"] not in allowed_roles:
                return render_template("error_403.html", current_user=user), 403
            return view_func(*args, **kwargs)
        return wrapper
    return decorator


def login_required_api(view_func):
    """Protege rotas de API: retorna 401 JSON se não autenticado."""
    @wraps(view_func)
    def wrapper(*args, **kwargs):
        if not get_current_user():
            return jsonify({
                "success": False,
                "error": "Sessão expirada ou não autenticada. Faça login novamente.",
                "auth_required": True,
            }), 401
        return view_func(*args, **kwargs)
    return wrapper


def role_required_api(*allowed_roles):
    """Protege rotas de API por papel (role). Retorna 403 JSON se não permitido."""
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(*args, **kwargs):
            user = get_current_user()
            if not user:
                return jsonify({
                    "success": False,
                    "error": "Sessão expirada ou não autenticada. Faça login novamente.",
                    "auth_required": True,
                }), 401
            if user["role"] not in allowed_roles:
                return jsonify({
                    "success": False,
                    "error": "Acesso negado: seu perfil não tem permissão para esta ação.",
                }), 403
            return view_func(*args, **kwargs)
        return wrapper
    return decorator


# --------------------------------------------------------------------------
# TRATAMENTO GLOBAL DE ERROS
# --------------------------------------------------------------------------

@app.errorhandler(404)
def handle_404(_error):
    if request.path.startswith("/api/"):
        return jsonify({"success": False, "error": "Recurso não encontrado."}), 404
    return render_template("error_404.html", current_user=get_current_user()), 404


@app.errorhandler(413)
def handle_413(_error):
    return jsonify({"success": False, "error": "Arquivo excede o tamanho máximo permitido (10MB)."}), 413


@app.errorhandler(500)
def handle_500(error):
    app.logger.error("Erro interno: %s", traceback.format_exc())
    if request.path.startswith("/api/"):
        return jsonify({"success": False, "error": "Erro interno do servidor."}), 500
    return render_template("error_500.html"), 500


@app.route("/sobre")
def page_sobre():
    """Página institucional: o que é o GateFlow e como funciona."""
    return render_template("sobre.html")


@app.route("/termos")
def page_termos():
    """Termos de Uso da plataforma."""
    return render_template("termos.html")


@app.route("/privacidade")
def page_privacidade():
    """Política de Privacidade da plataforma."""
    return render_template("privacidade.html")


# --------------------------------------------------------------------------
# PÁGINAS (FRONT-END SERVIDO PELO PRÓPRIO FLASK)
# --------------------------------------------------------------------------

# Quantos depoimentos vêm prontos (SSR) na landing page — o suficiente
# para preencher a seção sem round-trip extra no primeiro load. A lista
# COMPLETA só é buscada sob demanda, quando o visitante abre o modal
# "Ver todos os comentários" (ver GET /api/feedbacks) — mantém o load
# inicial da landing leve mesmo com centenas de depoimentos acumulados.
LANDING_FEEDBACKS_PREVIEW_COUNT = 6


@app.route("/")
def page_index():
    """Landing page pública: introdução ao sistema e botões de acesso."""
    try:
        initial_feedbacks = [
            _public_feedback_payload(f)
            for f in db.list_feedbacks(limit=LANDING_FEEDBACKS_PREVIEW_COUNT)
        ]
    except db.DatabaseError:
        initial_feedbacks = []
    return render_template(
        "index.html", current_user=get_current_user(), initial_feedbacks=initial_feedbacks
    )


def _redirect_url_for_role(role):
    """
    Centraliza o destino pós-login/cadastro de cada papel do GateFlow.
    Usado tanto pelo /login quanto pelo /organizador/cadastro, para as
    duas rotas nunca ficarem com lógicas de redirecionamento divergentes.
    """
    if role == "admin":
        return url_for("page_admin")
    if role == "organizador":
        return url_for("page_organizador_dashboard")
    if role == "cliente":
        # Área do Cliente Comprador é uma fase futura (Fase 4 do roteiro);
        # por ora manda para a landing em vez de uma página inexistente.
        return url_for("page_index")
    return url_for("page_checkin")  # porteiro (padrão)


@app.route("/login", methods=["GET", "POST"])
@limiter.limit("20 per minute")
def page_login():
    # Se já está logado, manda direto para o painel correspondente ao seu papel
    existing_user = get_current_user()
    if existing_user:
        return redirect(_redirect_url_for_role(existing_user["role"]))

    error = None
    if request.method == "POST":
        username = (request.form.get("username") or "").strip()
        password = request.form.get("password") or ""
        client_ip = request.remote_addr or "unknown"

        is_locked, seconds_remaining = login_attempt_tracker.is_locked(username, client_ip)
        if is_locked:
            minutes_remaining = max(1, seconds_remaining // 60)
            error = (
                f"Muitas tentativas de login falhas. Tente novamente em "
                f"aproximadamente {minutes_remaining} minuto(s)."
            )
            intent = request.args.get("intent", "porteiro")
            return render_template("login.html", error=error, intent=intent, active_panel="signin")

        user = db.get_user_by_username(username) if username else None
        if user and check_password_hash(user["password_hash"], password):
            login_attempt_tracker.register_success(username, client_ip)
            session.clear()
            session.permanent = True
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            session["username"] = user["username"]
            session["session_token"] = db.start_new_user_session(user["id"])
            return redirect(_redirect_url_for_role(user["role"]))

        login_attempt_tracker.register_failure(username, client_ip)
        error = "Usuário ou senha inválidos."

    # 'intent' só ajusta o texto exibido no formulário (não concede acesso —
    # o papel real vem do banco de dados após a autenticação). 'panel'
    # permite abrir a tela já com o painel de Cadastro ativo (ex: link
    # "Criar Conta" vindo de outra página).
    intent = request.args.get("intent", "porteiro")
    active_panel = "signup" if request.args.get("panel") == "signup" else "signin"
    return render_template("login.html", error=error, intent=intent, active_panel=active_panel)


@app.route("/organizador/cadastro", methods=["GET", "POST"])
def page_organizador_signup():
    """
    Cadastro autônomo (self-service) de novos Organizadores — Fase 1,
    item 2. Qualquer pessoa pode se cadastrar sem depender de um Super
    Admin criar a conta manualmente.
    """
    existing_user = get_current_user()
    if existing_user:
        return redirect(_redirect_url_for_role(existing_user["role"]))

    error = None
    if request.method == "POST":
        full_name = (request.form.get("full_name") or "").strip()
        email = (request.form.get("email") or "").strip()
        username = (request.form.get("username") or "").strip()
        phone = (request.form.get("phone") or "").strip()
        password = request.form.get("password") or ""
        password_confirm = request.form.get("password_confirm") or ""

        email_pattern = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

        if not full_name or not email or not username or not password:
            error = "Preencha nome completo, e-mail, usuário e senha."
        elif not email_pattern.match(email):
            error = "Informe um e-mail válido."
        elif len(password) < 8:
            error = "A senha deve ter pelo menos 8 caracteres."
        elif password != password_confirm:
            error = "As senhas não coincidem."
        elif db.username_exists(username):
            error = "Este nome de usuário já está em uso."
        elif db.email_exists(email):
            error = "Já existe uma conta com este e-mail."
        else:
            password_hash = generate_password_hash(password)
            user = db.create_user(
                username, password_hash, "organizador",
                full_name=full_name, email=email, phone=phone,
            )
            session.clear()
            session.permanent = True
            session["user_id"] = user["id"]
            session["role"] = user["role"]
            session["username"] = user["username"]
            session["session_token"] = db.start_new_user_session(user["id"])
            return redirect(url_for("page_organizador_dashboard"))

    # Ponto único de acesso (Fase 5 — redesign do login): tanto o GET
    # quanto um erro de validação no POST mantêm a pessoa na MESMA tela
    # de Login/Cadastro (painel deslizante), com o painel de cadastro
    # já ativo — nunca mais uma página separada por tipo de fluxo.
    return render_template("login.html", signup_error=error, active_panel="signup")


@app.route("/organizador")
@role_required_page("organizador")
def page_organizador_dashboard():
    """
    Painel do Organizador: criação de eventos (Módulo A/B) e listagem de
    todos os eventos da plataforma, com isolamento de gerenciamento.
    """
    return render_template(
        "organizador_dashboard.html",
        current_user=get_current_user(),
        max_porteiros=config.MAX_PORTEIROS_POR_ORGANIZADOR,
    )


@app.route("/organizador/eventos/<event_id>")
@role_required_page("organizador")
def page_organizador_event_detail(event_id):
    """Gerenciamento de um evento específico do Organizador — ownership obrigatória."""
    current_user = get_current_user()
    if not db.event_belongs_to_organizador(event_id, current_user["id"]):
        abort(404)
    event = db.get_event(event_id)
    event["status"] = event_status.compute_event_status(event.get("event_date"))
    return render_template("organizador_event_detail.html", current_user=current_user, event=event)


@app.route("/logout")
def page_logout():
    user_id = session.get("user_id")
    if user_id:
        db.start_new_user_session(user_id)  # invalida qualquer copia do cookie que ainda circule
    session.clear()
    return redirect(url_for("page_index"))


@app.route("/admin")
@role_required_page("admin")
def page_admin():
    """Painel Geral do Admin: lista TODOS os eventos cadastrados (Multi-Eventos).

    Os eventos sao carregados aqui e injetados no template (SSR) para que
    a lista ja apareca pronta no HTML entregue ao navegador, sem depender
    de um fetch('/api/events') subsequente nem do skeleton de loading.
    Se a query falhar por algum motivo, `initial_events` fica None e o
    admin_events.js cai de volta no fluxo antigo (fetch + skeleton).
    """
    current_user = get_current_user()
    try:
        initial_events = _list_events_for_user(current_user)
    except db.DatabaseError:
        initial_events = None
    return render_template(
        "admin_events.html", current_user=current_user, initial_events=initial_events
    )


@app.route("/admin/eventos/<event_id>")
@role_required_page("admin")
def page_admin_event_detail(event_id):
    """Gerenciamento de um evento específico (importar, exportar PDF, convidados)."""
    event = db.get_event(event_id)
    if not event:
        abort(404)
    event["status"] = event_status.compute_event_status(event.get("event_date"))
    return render_template(
        "admin_event_detail.html",
        current_user=get_current_user(),
        event=event,
    )


@app.route("/admin/qrcode-site")
@role_required_api("admin")
def admin_download_site_qrcode():
    """
    Gera, SOB DEMANDA e em memória (nunca salvo em disco -- importante
    em produção serverless, onde não há disco persistente garantido), o
    QR Code oficial do sistema: aponta para a URL pública configurada em
    GATEFLOW_PUBLIC_BASE_URL, para uso em material impresso/porta.

    Query string opcional: ?url=https://outra-url.com para gerar o QR
    de uma URL diferente da configurada (ex: testar antes de configurar
    a variável de ambiente definitiva).
    """
    from generate_site_qr import generate_site_qr

    target_url = (request.args.get("url") or config.PUBLIC_BASE_URL or request.host_url).strip()
    if not target_url:
        return jsonify({
            "success": False,
            "error": "Nenhuma URL configurada. Defina GATEFLOW_PUBLIC_BASE_URL ou passe ?url=...",
        }), 400

    buffer = io.BytesIO()
    try:
        # Reaproveita a mesma lógica do script generate_site_qr.py, mas
        # devolvendo bytes em memória em vez de gravar um arquivo .png.
        import qrcode
        from qrcode.constants import ERROR_CORRECT_H

        qr = qrcode.QRCode(version=None, error_correction=ERROR_CORRECT_H, box_size=12, border=4)
        qr.add_data(target_url)
        qr.make(fit=True)
        img = qr.make_image(fill_color="#000000", back_color="#FFFFFF").convert("RGB")
        img.save(buffer, format="PNG")
    except Exception:
        app.logger.exception("Falha ao gerar QR Code do site.")
        return jsonify({"success": False, "error": "Falha ao gerar o QR Code."}), 500

    buffer.seek(0)
    return send_file(buffer, mimetype="image/png", download_name="gateflow-qrcode-site.png")


@app.route("/admin/usuarios")
@role_required_page("admin")
def page_admin_users():
    """Gerenciamento de usuários: qualquer usuário com papel 'admin' pode criar novos perfis."""
    return render_template("admin_users.html", current_user=get_current_user())


@app.route("/checkin")
@role_required_page("admin", "porteiro")
def page_checkin():
    """Painel Geral do Porteiro: lista TODOS os eventos cadastrados (Multi-Eventos).

    Mesma logica de SSR aplicada em page_admin: os eventos ja vao prontos
    no HTML, entao o checkin_events.js nao precisa esperar por um
    fetch('/api/events') para pintar a lista real.
    """
    # Admin também pode acessar este painel, conforme especificado
    current_user = get_current_user()
    try:
        initial_events = _list_events_for_user(current_user)
    except db.DatabaseError:
        initial_events = None
    return render_template(
        "checkin_events.html", current_user=current_user, initial_events=initial_events
    )


@app.route("/checkin/<event_id>")
@role_required_page("admin", "porteiro")
def page_checkin_event(event_id):
    """Tela de check-in (busca + leitor de QR Code) para um evento específico."""
    event = db.get_event(event_id)
    if not event:
        abort(404)
    if not _porteiro_can_access_event(get_current_user(), event):
        abort(404)  # evento existe, mas fora do escopo deste porteiro — tratado como inexistente
    event["status"] = event_status.compute_event_status(event.get("event_date"))
    return render_template(
        "checkin.html",
        current_user=get_current_user(),
        event=event,
    )


# --------------------------------------------------------------------------
# API: EVENTOS
# --------------------------------------------------------------------------

def _list_events_for_user(current_user):
    """Retorna a lista de eventos (com status calculado e ordenados para
    exibição) visível para o usuário atual.

    Centraliza a regra de isolamento de porteiros do promotor (Fase 5,
    item C) num único lugar, usado tanto por /api/events quanto pela
    renderização inicial (SSR) de /admin e /checkin — garante que os
    dois caminhos nunca fiquem dessincronizados.
    """
    if current_user["role"] == "porteiro" and current_user.get("organizador_id"):
        events = db.list_events_owned_by(current_user["organizador_id"])
    else:
        events = db.list_events()
    event_status.attach_status(events)
    return event_status.sort_events_for_display(events)


@app.route("/api/events", methods=["GET"])
@role_required_api("admin", "porteiro")
def api_list_events():
    # Acessível a admin e porteiro: o porteiro precisa listar eventos
    # para escolher em qual fará check-in.
    try:
        events = _list_events_for_user(get_current_user())
        return jsonify({"success": True, "data": events})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/organizador/events", methods=["GET"])
@role_required_api("organizador")
def api_organizador_list_events():
    """
    Fase 2, item 3: agora retorna TODOS os eventos da plataforma (não só
    os do Organizador autenticado) — "para dar volume à plataforma".
    Cada evento vem marcado com `is_own` para o frontend desabilitar
    botões de gerenciamento nos que não pertencem a quem está olhando.

    IMPORTANTE (privacidade): mesmo aparecendo na listagem, eventos de
    OUTROS organizadores têm `total_guests`/`total_checked_in` ocultados
    (viram None) — a lista de convidados e o quanto ela já converteu em
    presença é informação sensível do dono do evento, não precisa vazar
    para concorrentes só porque o evento aparece na vitrine da plataforma.
    O isolamento de ACESSO (detalhe, edição, convidados) continua
    absoluto — isso aqui é só sobre o que aparece nesta listagem.
    """
    try:
        current_user = get_current_user()
        events = db.list_events()
        event_status.attach_status(events)
        events = event_status.sort_events_for_display(events)

        for event in events:
            is_own = event.get("organizador_id") == current_user["id"]
            event["is_own"] = is_own
            if not is_own:
                event["total_guests"] = None
                event["total_checked_in"] = None

        return jsonify({"success": True, "data": events})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


def _require_module_a_event(event):
    """
    Fase 5: segregacao estrita entre Modelo A (lista fechada/convidados) e
    Modelo B (bilheteira publica). Funcionalidades de convidados (cadastro
    manual, importacao de planilha, edicao, exclusao) sao EXCLUSIVAS do
    Modelo A -- um evento do Modelo B nao tem tabela de convidados fazendo
    sentido nenhum, e expor essas rotas para ele e um vazamento de UI/API
    entre os dois modelos de negocio. Retorna uma tupla (response, status)
    pronta para return se o evento nao for Modelo A, ou None se estiver tudo certo.
    """
    if event.get("event_module") != "A":
        return jsonify({
            "success": False,
            "error": "Esta funcionalidade e exclusiva de eventos do Modelo A (lista fechada de convidados).",
        }), 403
    return None


def _validate_module_b_capacity(payload):
    """
    Valida a lotação por setores do Módulo B. Retorna (vip, normal, total, error).
    Regra: se só "total" for informado, VIP vira 0 e Normal vira o total
    inteiro. Se VIP e Normal forem informados, a soma deve bater EXATAMENTE
    com o total informado.
    """
    def _parse_int(value):
        if value is None or value == "":
            return None
        try:
            parsed = int(value)
            return parsed if parsed >= 0 else None
        except (TypeError, ValueError):
            return None

    vip = _parse_int(payload.get("capacity_vip"))
    normal = _parse_int(payload.get("capacity_normal"))
    total = _parse_int(payload.get("capacity_total"))

    if total is None:
        return None, None, None, "A lotação total é obrigatória para eventos do Módulo B."

    if vip is None and normal is None:
        # Só a lotação total foi informada -> tudo vira "Normal"
        return 0, total, total, None

    vip = vip or 0
    normal = normal or 0
    if vip + normal != total:
        return None, None, None, (
            f"A soma de Vagas VIP ({vip}) + Vagas Normais ({normal}) = {vip + normal} "
            f"não bate com a Lotação Total informada ({total})."
        )

    return vip, normal, total, None


def _validate_module_b_prices(payload, capacity_vip, capacity_normal):
    """
    Valida os preços de venda ONLINE (Módulo B). Fase 5: a definição de
    preço por setor passa a ser OBRIGATÓRIA para qualquer setor que
    tenha vagas (capacity > 0) antes do evento poder ser publicado -- um
    evento do Modelo B sem preço configurado não tem como vender bilhete
    nenhum, então a plataforma não permite esse estado inconsistente
    nem por engano do organizador.
    Retorna (price_vip, price_normal, error).
    """
    def _parse_price(value):
        if value is None or value == "":
            return None
        try:
            parsed = float(value)
            return parsed if parsed >= 0 else None
        except (TypeError, ValueError):
            return None

    price_vip_raw = payload.get("price_vip")
    price_normal_raw = payload.get("price_normal")

    price_vip = _parse_price(price_vip_raw) if price_vip_raw not in (None, "") else 0
    price_normal = _parse_price(price_normal_raw) if price_normal_raw not in (None, "") else 0

    if price_vip is None:
        return None, None, "Preço VIP inválido."
    if price_normal is None:
        return None, None, "Preço Normal inválido."

    if (capacity_vip or 0) > 0 and price_vip <= 0:
        return None, None, "Defina o preço do ingresso VIP (maior que 0) antes de publicar este evento."
    if (capacity_normal or 0) > 0 and price_normal <= 0:
        return None, None, "Defina o preço do ingresso Normal (maior que 0) antes de publicar este evento."

    return price_vip, price_normal, None


@app.route("/api/organizador/events", methods=["POST"])
@role_required_api("organizador")
def api_organizador_create_event():
    """Criação de evento pelo Organizador — Fase 2, item 1. Suporta Módulo A e B."""
    try:
        current_user = get_current_user()
        payload = request.get_json(force=True, silent=True) or {}

        event_module = (payload.get("event_module") or "").strip().upper()
        name = (payload.get("name") or "").strip()
        organizer_display_name = (payload.get("organizer_display_name") or "").strip()
        location = (payload.get("location") or "").strip()
        event_date = (payload.get("event_date") or "").strip()
        description = (payload.get("description") or "").strip()

        if event_module not in db.VALID_EVENT_MODULES:
            return jsonify({"success": False, "error": "Selecione o Tipo de Evento (Módulo A ou B)."}), 400
        if not name or not organizer_display_name or not location or not event_date:
            return jsonify({
                "success": False,
                "error": "Nome do Evento, Nome do Organizador/Empresa, Localização e Data/Hora são obrigatórios.",
            }), 400

        if event_module == "A":
            event = db.create_event_full(
                current_user["id"], "A", name, organizer_display_name, location, event_date,
                description=description,
            )
        else:
            contact_phone = (payload.get("contact_phone") or "").strip()
            if not contact_phone:
                return jsonify({
                    "success": False,
                    "error": "Contato Telefónico/WhatsApp é obrigatório para eventos do Módulo B.",
                }), 400

            vip, normal, total, capacity_error = _validate_module_b_capacity(payload)
            if capacity_error:
                return jsonify({"success": False, "error": capacity_error}), 400

            price_vip, price_normal, price_error = _validate_module_b_prices(payload, vip, normal)
            if price_error:
                return jsonify({"success": False, "error": price_error}), 400

            event = db.create_event_full(
                current_user["id"], "B", name, organizer_display_name, location, event_date,
                description=description, contact_phone=contact_phone,
                capacity_vip=vip, capacity_normal=normal, capacity_total=total,
                price_vip=price_vip, price_normal=price_normal,
            )

        event["status"] = event_status.compute_event_status(event.get("event_date"))
        return jsonify({"success": True, "data": event}), 201
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except Exception as exc:  # noqa: BLE001
        app.logger.error("Erro inesperado ao criar evento (organizador): %s", traceback.format_exc())
        return jsonify({"success": False, "error": f"Erro ao criar evento: {exc}"}), 500


@app.route("/api/organizador/events/<event_id>", methods=["GET"])
@role_required_api("organizador")
def api_organizador_get_event(event_id):
    """Detalhe de um evento — SOMENTE se pertencer ao Organizador autenticado."""
    try:
        current_user = get_current_user()
        if not db.event_belongs_to_organizador(event_id, current_user["id"]):
            return jsonify({"success": False, "error": "Evento não encontrado ou não pertence a você."}), 404

        event = db.get_event(event_id)
        event["stats"] = db.get_event_stats(event_id)
        event["status"] = event_status.compute_event_status(event.get("event_date"))
        return jsonify({"success": True, "data": event})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/organizador/events/<event_id>", methods=["PUT"])
@role_required_api("organizador")
def api_organizador_update_event(event_id):
    """Edição de evento pelo Organizador — ownership + bloqueio de evento encerrado."""
    try:
        current_user = get_current_user()
        current_event = db.get_event(event_id)
        if not current_event or current_event.get("organizador_id") != current_user["id"]:
            return jsonify({"success": False, "error": "Evento não encontrado ou não pertence a você."}), 404

        if event_status.is_past(current_event.get("event_date")):
            return jsonify({
                "success": False,
                "error": "Este evento já foi encerrado. Edição bloqueada, apenas relatórios continuam disponíveis.",
                "event_ended": True,
            }), 403

        payload = request.get_json(force=True, silent=True) or {}
        update_fields = {
            "name": (payload.get("name") or "").strip() or None,
            "organizer_display_name": (payload.get("organizer_display_name") or "").strip() or None,
            "location": (payload.get("location") or "").strip() or None,
            "event_date": (payload.get("event_date") or "").strip() or None,
            "description": payload.get("description", None),
        }

        if current_event["event_module"] == "B" and any(
            k in payload for k in ("capacity_vip", "capacity_normal", "capacity_total")
        ):
            vip, normal, total, capacity_error = _validate_module_b_capacity(payload)
            if capacity_error:
                return jsonify({"success": False, "error": capacity_error}), 400
            update_fields.update(capacity_vip=vip, capacity_normal=normal, capacity_total=total)

        if current_event["event_module"] == "B" and any(k in payload for k in ("price_vip", "price_normal")):
            effective_vip = update_fields.get("capacity_vip", current_event.get("capacity_vip"))
            effective_normal = update_fields.get("capacity_normal", current_event.get("capacity_normal"))
            price_payload = {
                "price_vip": payload.get("price_vip", current_event.get("price_vip")),
                "price_normal": payload.get("price_normal", current_event.get("price_normal")),
            }
            price_vip, price_normal, price_error = _validate_module_b_prices(
                price_payload, effective_vip, effective_normal
            )
            if price_error:
                return jsonify({"success": False, "error": price_error}), 400
            update_fields.update(price_vip=price_vip, price_normal=price_normal)

        if "contact_phone" in payload:
            update_fields["contact_phone"] = (payload.get("contact_phone") or "").strip() or None

        event, owned = db.update_event_by_organizador(event_id, current_user["id"], **update_fields)
        if not owned:
            return jsonify({"success": False, "error": "Evento não encontrado ou não pertence a você."}), 404

        event["status"] = event_status.compute_event_status(event.get("event_date"))
        return jsonify({"success": True, "data": event})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


def _get_owned_module_a_event(event_id, organizador_id):
    """
    Helper compartilhado pelas rotas de gestão de convidados do
    Organizador: confirma posse do evento e que é do Módulo A. Retorna
    (event, error_response). `error_response` é None quando tudo está
    OK; caso contrário, já é a tupla (jsonify(...), status_code) pronta
    para a rota retornar direto.

    Antes havia aqui um paywall (`is_paid`) que liberava o acesso só
    após confirmação de pagamento do evento — removido junto com todo o
    módulo de vendas/monetização; a gestão de convidados agora está
    disponível assim que o evento é criado.
    """
    event = db.get_event(event_id)
    if not event or event.get("organizador_id") != organizador_id:
        return None, (jsonify({"success": False, "error": "Evento não encontrado ou não pertence a você."}), 404)

    if event.get("event_module") != "A":
        return None, (jsonify({
            "success": False,
            "error": "Gestão de lista de convidados só está disponível para eventos do Módulo A.",
        }), 400)

    return event, None


@app.route("/api/organizador/events/<event_id>/guests", methods=["GET"])
@role_required_api("organizador")
def api_organizador_list_guests(event_id):
    current_user = get_current_user()
    event, error = _get_owned_module_a_event(event_id, current_user["id"])
    if error:
        return error
    try:
        search = request.args.get("search", "").strip()
        guests = db.list_guests(event_id, search=search or None)
        stats = db.get_event_stats(event_id)
        return jsonify({"success": True, "data": guests, "stats": stats})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/organizador/events/<event_id>/guests", methods=["POST"])
@role_required_api("organizador")
def api_organizador_create_guest_manual(event_id):
    current_user = get_current_user()
    event, error = _get_owned_module_a_event(event_id, current_user["id"])
    if error:
        return error
    try:
        payload = request.get_json(force=True, silent=True) or {}
        full_name = (payload.get("full_name") or "").strip()
        email = (payload.get("email") or "").strip()
        phone = (payload.get("phone") or "").strip()
        role = (payload.get("role") or "").strip()
        table_number = (payload.get("table_number") or "").strip()

        if not full_name:
            return jsonify({"success": False, "error": "O nome completo é obrigatório."}), 400

        guest, is_duplicate = db.create_single_guest(event_id, full_name, email, phone, role, table_number)
        if is_duplicate:
            return jsonify({
                "success": False,
                "error": f"Já existe um convidado igual neste evento: {guest['full_name']}.",
                "data": guest,
            }), 409

        try:
            qr_path = generate_guest_qrcode(event_id, guest["id"])
            db.update_guest_qr_path(guest["id"], qr_path)
            guest["qr_code_path"] = qr_path
        except QRCodeGenerationError as exc:
            app.logger.error("Falha ao gerar QR Code do convidado (organizador): %s", exc)

        return jsonify({"success": True, "data": guest}), 201
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except Exception as exc:  # noqa: BLE001
        app.logger.error("Erro inesperado no cadastro manual (organizador): %s", traceback.format_exc())
        return jsonify({"success": False, "error": f"Erro ao cadastrar convidado: {exc}"}), 500


@app.route("/api/organizador/events/<event_id>/import", methods=["POST"])
@role_required_api("organizador")
def api_organizador_import_guests(event_id):
    current_user = get_current_user()
    event, error = _get_owned_module_a_event(event_id, current_user["id"])
    if error:
        return error

    if "file" not in request.files:
        return jsonify({"success": False, "error": "Nenhum arquivo enviado."}), 400
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "error": "Nome de arquivo inválido."}), 400

    filename = secure_filename(file.filename)
    _, ext = os.path.splitext(filename.lower())
    if ext not in config.ALLOWED_UPLOAD_EXTENSIONS:
        return jsonify({"success": False, "error": "Apenas arquivos .xlsx são aceitos."}), 400

    temp_path = os.path.join(config.UPLOADS_DIR, filename)
    try:
        file.save(temp_path)
        guests_data, warnings = _parse_guests_xlsx(temp_path)

        if not guests_data:
            return jsonify({"success": False, "error": "Nenhum convidado válido encontrado na planilha."}), 400

        created_guests, skipped_count = db.import_guests_with_dedup(event_id, guests_data)

        generation_errors = []
        for guest in created_guests:
            try:
                qr_path = generate_guest_qrcode(event_id, guest["id"])
                db.update_guest_qr_path(guest["id"], qr_path)
            except QRCodeGenerationError as exc:
                generation_errors.append(str(exc))

        summary = (
            f"{len(created_guests)} convidado(s) adicionado(s), "
            f"{skipped_count} pulado(s) por já existirem neste evento."
        )
        return jsonify({
            "success": True,
            "data": {
                "imported_count": len(created_guests),
                "skipped_count": skipped_count,
                "summary": summary,
                "warnings": warnings,
                "qr_generation_errors": generation_errors,
            },
        }), 201
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:  # noqa: BLE001
        app.logger.error("Erro na importação (organizador): %s", traceback.format_exc())
        return jsonify({"success": False, "error": f"Erro ao processar a planilha: {exc}"}), 500
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass


@app.route("/api/organizador/events/<event_id>/logo", methods=["POST"])
@role_required_api("organizador")
def api_organizador_upload_logo(event_id):
    """
    Upload do logo de um evento (Módulo B, campo opcional). Validação em
    duas camadas: (1) extensão do arquivo, (2) o arquivo REALMENTE abre
    como imagem válida via Pillow — protege contra um arquivo malicioso
    disfarçado com extensão .png/.jpg.
    """
    current_user = get_current_user()
    event = db.get_event(event_id)
    if not event or event.get("organizador_id") != current_user["id"]:
        return jsonify({"success": False, "error": "Evento não encontrado ou não pertence a você."}), 404
    if event.get("event_module") != "B":
        return jsonify({"success": False, "error": "Logo só está disponível para eventos do Módulo B."}), 400
    if event_status.is_past(event.get("event_date")):
        return jsonify({"success": False, "error": "Este evento já foi encerrado.", "event_ended": True}), 403

    if "logo" not in request.files:
        return jsonify({"success": False, "error": "Nenhum arquivo de logo enviado."}), 400
    file = request.files["logo"]
    if file.filename == "":
        return jsonify({"success": False, "error": "Nome de arquivo inválido."}), 400

    filename = secure_filename(file.filename)
    _, ext = os.path.splitext(filename.lower())
    if ext not in config.ALLOWED_LOGO_EXTENSIONS:
        return jsonify({
            "success": False,
            "error": f"Formato não permitido. Use: {', '.join(config.ALLOWED_LOGO_EXTENSIONS)}.",
        }), 400

    file.seek(0, os.SEEK_END)
    file_size = file.tell()
    file.seek(0)
    if file_size > config.MAX_LOGO_SIZE_BYTES:
        max_mb = config.MAX_LOGO_SIZE_BYTES / (1024 * 1024)
        return jsonify({"success": False, "error": f"Imagem excede o tamanho máximo permitido ({max_mb:.0f}MB)."}), 400

    try:
        # Validação REAL do conteúdo: tenta abrir como imagem antes de
        # aceitar. Se falhar, não é uma imagem de verdade (ou está
        # corrompida), independente do que a extensão do arquivo diz.
        image = Image.open(file.stream)
        image.verify()
        file.stream.seek(0)
    except (UnidentifiedImageError, OSError):
        return jsonify({"success": False, "error": "O arquivo enviado não é uma imagem válida."}), 400

    try:
        # Guardamos o logo como um "data URL" (base64) dentro da própria
        # coluna `logo_path` (TEXT) em vez de salvar um arquivo em disco.
        # Isso é o que torna o logo de fato PERSISTENTE em produção
        # serverless: um arquivo em disco sobreviveria só até o fim
        # desta invocação (a Vercel não garante disco compartilhado nem
        # persistente entre requisições), enquanto o banco (Neon) é a
        # nossa única fonte de persistência real. Como logos são pequenos
        # (limite de 2MB, ver config.MAX_LOGO_SIZE_BYTES) o overhead de
        # ~33% do base64 é insignificante.
        file.stream.seek(0)
        raw_bytes = file.stream.read()
        content_type = Image.MIME.get(image.format, "image/png") if image.format else "image/png"
        logo_data_url = f"data:{content_type};base64,{base64.b64encode(raw_bytes).decode('ascii')}"

        updated_event, _ = db.update_event_by_organizador(event_id, current_user["id"], logo_path=logo_data_url)
        return jsonify({"success": True, "data": updated_event}), 201
    except Exception as exc:  # noqa: BLE001
        app.logger.error("Erro ao salvar logo: %s", traceback.format_exc())
        return jsonify({"success": False, "error": f"Erro ao salvar o logo: {exc}"}), 500


@app.route("/api/events/<event_id>/logo", methods=["GET"])
def api_get_event_logo(event_id):
    """Serve a imagem do logo de um evento. Pública (usada na vitrine de eventos)."""
    event = db.get_event(event_id)
    logo_value = event.get("logo_path") if event else None
    if not logo_value:
        abort(404)

    if logo_value.startswith("data:"):
        # Formato novo: "data:<content-type>;base64,<...>" guardado direto
        # no banco (ver api_organizador_upload_logo acima) -- funciona
        # igual em qualquer ambiente, sem depender de disco.
        try:
            header, b64_content = logo_value.split(",", 1)
            content_type = header.split(";")[0].replace("data:", "") or "image/png"
            image_bytes = base64.b64decode(b64_content)
        except (ValueError, base64.binascii.Error):
            abort(404)
        return send_file(io.BytesIO(image_bytes), mimetype=content_type)

    # Formato antigo (instalações locais anteriores a esta mudança):
    # `logo_path` ainda é um caminho de arquivo em disco.
    if not os.path.exists(logo_value):
        abort(404)
    return send_file(logo_value)


# ============================================================================
# FASE 6 — DEPOIMENTOS/AVALIAÇÕES PÚBLICAS (LANDING PAGE)
# ============================================================================

_FEEDBACK_EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
FEEDBACK_COMMENT_MAX_LENGTH = 500
FEEDBACK_NAME_MAX_LENGTH = 80


def _public_feedback_payload(feedback):
    """
    Campos do depoimento seguros para expor publicamente. O e-mail NUNCA
    é devolvido em texto puro — só o hash MD5 (exigido pelo protocolo do
    Gravatar) embutido já pronto na avatar_url, para o frontend não
    precisar calcular hash nenhum nem ver o e-mail de quem comentou.
    """
    email_hash = hashlib.md5(feedback["email"].strip().lower().encode("utf-8")).hexdigest()
    return {
        "id": feedback["id"],
        "name": feedback["name"],
        "rating": feedback["rating"],
        "comment": feedback["comment"],
        "created_at": feedback["created_at"],
        # d=identicon: gera um avatar geométrico consistente quando o
        # e-mail não tem foto cadastrada no Gravatar -- nunca cai numa
        # imagem quebrada. s=80: pede já o tamanho final ao Gravatar
        # (mais leve que baixar uma imagem grande e encolher no CSS).
        "avatar_url": f"https://www.gravatar.com/avatar/{email_hash}?d=identicon&s=80",
    }


@app.route("/api/feedbacks", methods=["GET"])
@limiter.limit("60 per minute")
def api_list_feedbacks():
    """Lista os depoimentos JÁ APROVADOS — usado pelo modal 'Ver todos os comentários' da landing page."""
    try:
        feedbacks = db.list_feedbacks()
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    return jsonify({"success": True, "data": [_public_feedback_payload(f) for f in feedbacks]})


@app.route("/api/feedbacks", methods=["POST"])
@limiter.limit("5 per minute")
def api_create_feedback():
    """
    Recebe um novo depoimento do formulário público da landing page.
    Sem autenticação (qualquer visitante pode enviar), por isso o rate
    limit é mais apertado que o das outras rotas públicas e toda entrada
    é validada e limitada em tamanho antes de tocar no banco.

    Fica com status='pending' (db.create_feedback) e só aparece
    publicamente depois de um admin aprovar em /admin/depoimentos.
    """
    payload = request.get_json(force=True, silent=True) or {}
    name = (payload.get("name") or "").strip()
    email = (payload.get("email") or "").strip()
    comment = (payload.get("comment") or "").strip()
    try:
        rating = int(payload.get("rating"))
    except (TypeError, ValueError):
        rating = 0

    if not name or len(name) > FEEDBACK_NAME_MAX_LENGTH:
        return jsonify({
            "success": False,
            "error": f"Indique o seu nome (até {FEEDBACK_NAME_MAX_LENGTH} caracteres).",
        }), 400
    if not email or not _FEEDBACK_EMAIL_PATTERN.match(email) or len(email) > 150:
        return jsonify({"success": False, "error": "Indique um e-mail válido."}), 400
    if rating < 1 or rating > 5:
        return jsonify({"success": False, "error": "A nota deve ser de 1 a 5 estrelas."}), 400
    if not comment:
        return jsonify({"success": False, "error": "Escreva o seu comentário."}), 400
    if len(comment) > FEEDBACK_COMMENT_MAX_LENGTH:
        return jsonify({
            "success": False,
            "error": f"Comentário muito longo (máx. {FEEDBACK_COMMENT_MAX_LENGTH} caracteres).",
        }), 400

    try:
        feedback = db.create_feedback(name, email, rating, comment)
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500

    return jsonify({
        "success": True,
        "data": _public_feedback_payload(feedback),
        "message": "Obrigado pelo seu depoimento! Ele será exibido na página assim que for verificado pela nossa equipa.",
    }), 201


# --- Moderação (admin) ---

@app.route("/admin/depoimentos")
@role_required_page("admin")
def page_admin_feedbacks():
    """
    Fila de moderação de depoimentos: só aprovados aparecem na landing
    (ver db.list_feedbacks), então todo depoimento novo passa por aqui
    primeiro. A lista pendente já vem pronta via SSR (mesmo padrão de
    /admin e /checkin) — sem fetch extra no primeiro paint.
    """
    try:
        pending_feedbacks = db.list_pending_feedbacks()
    except db.DatabaseError:
        pending_feedbacks = None
    return render_template(
        "admin_feedbacks.html",
        current_user=get_current_user(),
        initial_pending_feedbacks=pending_feedbacks,
    )


@app.route("/api/admin/feedbacks/pending", methods=["GET"])
@role_required_api("admin")
def api_admin_list_pending_feedbacks():
    try:
        feedbacks = db.list_pending_feedbacks()
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    return jsonify({"success": True, "data": feedbacks})


@app.route("/api/admin/feedbacks/<feedback_id>/approve", methods=["PUT"])
@role_required_api("admin")
def api_admin_approve_feedback(feedback_id):
    try:
        updated = db.approve_feedback(feedback_id)
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    if not updated:
        return jsonify({"success": False, "error": "Depoimento não encontrado."}), 404
    return jsonify({"success": True})


@app.route("/api/admin/feedbacks/<feedback_id>/reject", methods=["PUT"])
@role_required_api("admin")
def api_admin_reject_feedback(feedback_id):
    try:
        updated = db.reject_feedback(feedback_id)
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    if not updated:
        return jsonify({"success": False, "error": "Depoimento não encontrado."}), 404
    return jsonify({"success": True})


# ============================================================================
# FASE 5 - PILAR 2: CONTROLO DE ACESSO EM TEMPO REAL (PORTARIA)
# ============================================================================

@app.route("/api/checkin/events/<event_id>/live-stats", methods=["GET"])
@role_required_api("admin", "organizador", "porteiro")
def api_checkin_live_stats(event_id):
    """Contadores em tempo real da portaria (vendidos/convidados vs entradas) e distribuicao de entradas por hora, para o dashboard de check-in."""
    current_user = get_current_user()
    if current_user["role"] == "porteiro":
        event = db.get_event(event_id)
        if not _porteiro_can_access_event(current_user, event):
            return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    elif current_user["role"] == "organizador":
        if not db.event_belongs_to_organizador(event_id, current_user["id"]):
            return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    stats = db.get_event_live_checkin_stats(event_id)
    if stats is None:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    return jsonify({"success": True, "data": stats})


# ============================================================================
# FASE 5 - PILAR 3: GESTAO DE CORTESIAS/VIP (MODELO A)
# ============================================================================

def _issue_courtesy(event_id, payload):
    full_name = (payload.get("full_name") or "").strip()
    email = (payload.get("email") or "").strip() or None
    phone = (payload.get("phone") or "").strip() or None
    table_number = (payload.get("table_number") or "").strip()

    guest, error = db.issue_courtesy_guest(event_id, full_name, email, phone, table_number)
    if error:
        return jsonify({"success": False, "error": error}), 409

    try:
        qr_path = generate_guest_qrcode(event_id, guest["id"])
        db.update_guest_qr_path(guest["id"], qr_path)
        guest["qr_code_path"] = qr_path
    except QRCodeGenerationError as exc:
        return jsonify({"success": False, "error": f"Convidado criado, mas falhou ao gerar QR Code: {exc}"}), 500

    if guest.get("email"):
        try:
            event = db.get_event(event_id)
            pdf_path = generate_invites_pdf(event, [guest])
            email_service.send_courtesy_invite_email(guest["email"], guest["full_name"], event["name"], pdf_path)
        except Exception as exc:  # noqa: BLE001 - falha de e-mail nunca invalida a emissao do convite
            app.logger.warning("Falha ao enviar convite de cortesia por e-mail: %s", exc)

    return jsonify({"success": True, "data": guest}), 201


# ============================================================================
# CONVITES / RSVP
# ============================================================================

def _build_rsvp_url(token):
    base_url = config.PUBLIC_BASE_URL or request.host_url.rstrip("/")
    return f"{base_url}{url_for('page_rsvp', token=token)}"


def _build_guest_qr_url(token):
    base_url = config.PUBLIC_BASE_URL or request.host_url.rstrip("/")
    return f"{base_url}{url_for('rsvp_qrcode', token=token)}"


def _dispatch_guest_invite(event, guest):
    """
    Envia o convite (com o link pessoal de RSVP) por WhatsApp; se falhar
    ou não houver telefone, cai para e-mail. Nunca lança -- melhor
    esforço, como o resto das notificações do sistema.
    """
    token = db.ensure_rsvp_token(guest["id"])
    rsvp_url = _build_rsvp_url(token)

    sent_ok = False
    phone = normalize_msisdn(guest.get("phone"))
    if phone:
        try:
            sent_ok, _err = whatsapp_service.send_rsvp_invite_whatsapp(phone, guest["full_name"], event["name"], rsvp_url)
        except Exception as exc:  # noqa: BLE001
            app.logger.warning("Falha ao enviar convite WhatsApp para %s: %s", guest["id"], exc)
        if not sent_ok:
            try:
                sms_service.send_rsvp_invite_sms(phone, guest["full_name"], event["name"], rsvp_url)
            except Exception as exc:  # noqa: BLE001
                app.logger.warning("Falha ao enviar convite SMS para %s: %s", guest["id"], exc)

    if guest.get("email"):
        try:
            email_service.send_rsvp_invite_email(guest["email"], guest["full_name"], event["name"], rsvp_url)
        except Exception as exc:  # noqa: BLE001
            app.logger.warning("Falha ao enviar convite por e-mail para %s: %s", guest["id"], exc)

    db.mark_guest_invite_sent(guest["id"])


def _dispatch_guest_reminder(event, guest):
    """Envia o lembrete do dia do evento (horário + localização + QR Code) -- WhatsApp, senão e-mail."""
    token = db.ensure_rsvp_token(guest["id"])
    qr_url = _build_guest_qr_url(token)

    try:
        event_dt = datetime.fromisoformat(event["event_date"])
        date_label = event_dt.strftime("%d/%m/%Y às %H:%M")
    except (ValueError, TypeError):
        date_label = event.get("event_date", "")

    maps_url = event.get("location_maps_url") or ""

    sent_ok = False
    phone = normalize_msisdn(guest.get("phone"))
    if phone:
        try:
            sent_ok, _err = whatsapp_service.send_rsvp_reminder_whatsapp(
                phone, guest["full_name"], event["name"], date_label, maps_url, qr_url
            )
        except Exception as exc:  # noqa: BLE001
            app.logger.warning("Falha ao enviar lembrete WhatsApp para %s: %s", guest["id"], exc)
        if not sent_ok:
            try:
                sms_service.send_rsvp_reminder_sms(phone, guest["full_name"], event["name"], date_label, qr_url)
            except Exception as exc:  # noqa: BLE001
                app.logger.warning("Falha ao enviar lembrete SMS para %s: %s", guest["id"], exc)

    if guest.get("email"):
        try:
            qr_path = guest.get("qr_code_path")
            if not qr_path or not os.path.exists(qr_path):
                qr_path = generate_guest_qrcode(event["id"], guest["id"])
            email_service.send_rsvp_reminder_email(guest["email"], guest["full_name"], event["name"], date_label, maps_url, qr_path)
        except Exception as exc:  # noqa: BLE001
            app.logger.warning("Falha ao enviar lembrete por e-mail para %s: %s", guest["id"], exc)

    db.mark_guest_reminder_sent(guest["id"])


@app.route("/api/organizador/events/<event_id>/guests/<guest_id>/send-invite", methods=["POST"])
@role_required_api("organizador")
@limiter.limit("60 per minute")
def api_organizador_send_guest_invite(event_id, guest_id):
    current_user = get_current_user()
    event, error = _get_owned_module_a_event(event_id, current_user["id"])
    if error:
        return error
    guest = db.get_guest_by_event(event_id, guest_id)
    if not guest:
        return jsonify({"success": False, "error": "Convidado não encontrado."}), 404
    _dispatch_guest_invite(event, guest)
    return jsonify({"success": True, "data": db.get_guest(guest_id)})


@app.route("/api/organizador/events/<event_id>/guests/send-invites", methods=["POST"])
@role_required_api("organizador")
@limiter.limit("10 per minute")
def api_organizador_send_invites_bulk(event_id):
    """Dispara o convite para todos os convidados que ainda não o receberam."""
    current_user = get_current_user()
    event, error = _get_owned_module_a_event(event_id, current_user["id"])
    if error:
        return error
    pending = db.list_guests_pending_invite(event_id)
    for guest in pending:
        _dispatch_guest_invite(event, guest)
    return jsonify({"success": True, "data": {"sent": len(pending)}})


@app.route("/api/organizador/events/<event_id>/guests/<guest_id>/send-reminder", methods=["POST"])
@role_required_api("organizador")
@limiter.limit("60 per minute")
def api_organizador_send_guest_reminder(event_id, guest_id):
    current_user = get_current_user()
    event, error = _get_owned_module_a_event(event_id, current_user["id"])
    if error:
        return error
    guest = db.get_guest_by_event(event_id, guest_id)
    if not guest:
        return jsonify({"success": False, "error": "Convidado não encontrado."}), 404
    if guest["rsvp_status"] != "confirmed":
        return jsonify({"success": False, "error": "Só é possível enviar lembrete a convidados que confirmaram presença."}), 400
    _dispatch_guest_reminder(event, guest)
    return jsonify({"success": True, "data": db.get_guest(guest_id)})


@app.route("/api/organizador/events/<event_id>/guests/send-reminders", methods=["POST"])
@role_required_api("organizador")
@limiter.limit("10 per minute")
def api_organizador_send_reminders_bulk(event_id):
    """Dispara o lembrete (QR + horário + mapa) para todos os confirmados que ainda não o receberam."""
    current_user = get_current_user()
    event, error = _get_owned_module_a_event(event_id, current_user["id"])
    if error:
        return error
    pending = db.list_guests_pending_reminder(event_id)
    for guest in pending:
        _dispatch_guest_reminder(event, guest)
    return jsonify({"success": True, "data": {"sent": len(pending)}})


@app.route("/api/organizador/events/<event_id>/guests/<guest_id>/companions", methods=["PUT"])
@role_required_api("organizador")
def api_organizador_set_guest_companions(event_id, guest_id):
    current_user = get_current_user()
    event, error = _get_owned_module_a_event(event_id, current_user["id"])
    if error:
        return error
    payload = request.get_json(force=True, silent=True) or {}
    try:
        companions_allowed = max(0, int(payload.get("companions_allowed", 0)))
    except (TypeError, ValueError):
        return jsonify({"success": False, "error": "companions_allowed deve ser um número inteiro."}), 400
    guest = db.set_guest_companions_allowed(event_id, guest_id, companions_allowed)
    if not guest:
        return jsonify({"success": False, "error": "Convidado não encontrado."}), 404
    return jsonify({"success": True, "data": guest})


@app.route("/api/organizador/events/<event_id>/rsvp-settings", methods=["PUT"])
@role_required_api("organizador")
def api_organizador_set_rsvp_settings(event_id):
    current_user = get_current_user()
    event, error = _get_owned_module_a_event(event_id, current_user["id"])
    if error:
        return error
    payload = request.get_json(force=True, silent=True) or {}
    deadline_days = payload.get("rsvp_deadline_days")
    try:
        deadline_days = int(deadline_days) if deadline_days not in (None, "") else None
    except (TypeError, ValueError):
        return jsonify({"success": False, "error": "rsvp_deadline_days deve ser um número inteiro de dias."}), 400
    maps_url = (payload.get("location_maps_url") or "").strip() or None
    updated = db.set_event_rsvp_settings(event_id, deadline_days, maps_url)
    return jsonify({"success": True, "data": updated})


@app.route("/api/organizador/events/<event_id>/rsvp-summary", methods=["GET"])
@role_required_api("organizador", "admin")
def api_organizador_rsvp_summary(event_id):
    """Resumo agregado (confirmados/recusados/pendentes/acompanhantes) -- usado pelo dashboard em tempo real (polling)."""
    current_user = get_current_user()
    if current_user["role"] == "organizador":
        event, error = _get_owned_module_a_event(event_id, current_user["id"])
        if error:
            return error
    try:
        summary = db.get_rsvp_summary(event_id)
        return jsonify({"success": True, "data": summary})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


# --- Página pública do convidado (/rsvp/<token>) ---

@app.route("/rsvp/<token>")
def page_rsvp(token):
    """Landing page pessoal do convidado: confirma presença (e acompanhantes) ou vê que o link expirou."""
    guest = db.get_guest_by_rsvp_token(token)
    if not guest:
        abort(404)

    deadline = db.rsvp_deadline_for_event(guest)
    expired = bool(deadline and datetime.utcnow() > deadline)

    return render_template("rsvp.html", guest=guest, expired=expired)


@app.route("/rsvp/<token>/qrcode.png")
def rsvp_qrcode(token):
    """QR Code de acesso do convidado -- link público (o token já é o segredo), usado nos lembretes por WhatsApp/SMS/e-mail."""
    guest = db.get_guest_by_rsvp_token(token)
    if not guest:
        abort(404)
    qr_path = guest.get("qr_code_path")
    if not qr_path or not os.path.exists(qr_path):
        try:
            qr_path = generate_guest_qrcode(guest["event_id"], guest["id"])
            db.update_guest_qr_path(guest["id"], qr_path)
        except QRCodeGenerationError:
            abort(404)
    return send_file(qr_path, mimetype="image/png")


@app.route("/api/rsvp/<token>/respond", methods=["POST"])
@limiter.limit("20 per minute")
def api_rsvp_respond(token):
    guest = db.get_guest_by_rsvp_token(token)
    if not guest:
        return jsonify({"success": False, "error": "Convite não encontrado."}), 404

    deadline = db.rsvp_deadline_for_event(guest)
    if deadline and datetime.utcnow() > deadline:
        return jsonify({"success": False, "error": "O prazo para confirmar presença neste evento já terminou."}), 410

    payload = request.get_json(force=True, silent=True) or {}
    response = payload.get("response")
    if response not in ("confirmed", "declined"):
        return jsonify({"success": False, "error": "Resposta inválida."}), 400

    try:
        companions_confirmed = max(0, min(int(payload.get("companions_confirmed", 0)), guest.get("companions_allowed") or 0))
    except (TypeError, ValueError):
        companions_confirmed = 0
    if response == "declined":
        companions_confirmed = 0

    updated = db.submit_rsvp_response(guest["id"], response, companions_confirmed, source="web")
    return jsonify({"success": True, "data": updated})


@app.route("/webhooks/whatsapp", methods=["GET", "POST"])
@limiter.limit("120 per minute")
def webhook_whatsapp_inbound():
    """
    Receptor de respostas inbound do WhatsApp ("1" = confirmar, "2" =
    recusar). Compatível com o formato de callback da WhatsApp Business
    Cloud API (Meta) -- se o provider configurado em WHATSAPP_PROVIDER
    for outro, ajuste o parsing do payload abaixo para o formato dele.

    GET: usado pela Meta para verificar o endpoint na configuração do
    webhook (responde ao hub.challenge). POST: mensagens recebidas.
    """
    if request.method == "GET":
        challenge = request.args.get("hub.challenge", "")
        return challenge, 200

    payload = request.get_json(force=True, silent=True) or {}
    try:
        entry = (payload.get("entry") or [{}])[0]
        change = (entry.get("changes") or [{}])[0]
        messages = change.get("value", {}).get("messages") or []
    except (IndexError, AttributeError):
        messages = []

    for msg in messages:
        from_msisdn = normalize_msisdn(msg.get("from", ""))
        text = (msg.get("text", {}) or {}).get("body", "").strip()
        if not from_msisdn or text not in ("1", "2"):
            continue

        guest = db.get_guest_by_phone_pending_rsvp(from_msisdn)
        if not guest:
            continue

        deadline = db.rsvp_deadline_for_event(guest)
        if deadline and datetime.utcnow() > deadline:
            continue

        response = "confirmed" if text == "1" else "declined"
        db.submit_rsvp_response(guest["id"], response, companions_confirmed=0, source="whatsapp")

    return jsonify({"success": True}), 200

@app.route("/api/organizador/events/<event_id>/guests/courtesy", methods=["POST"])
@role_required_api("organizador")
def api_organizador_issue_courtesy(event_id):
    """Emite um convidado Cortesia/VIP (Modelo A), respeitando a cota do evento, e envia o convite em PDF por e-mail se houver endereco."""
    current_user = get_current_user()
    if not db.event_belongs_to_organizador(event_id, current_user["id"]):
        return jsonify({"success": False, "error": "Evento não encontrado ou não pertence a você."}), 404
    payload = request.get_json(force=True, silent=True) or {}
    return _issue_courtesy(event_id, payload)


@app.route("/api/events/<event_id>/guests/courtesy", methods=["POST"])
@role_required_api("admin")
def api_admin_issue_courtesy(event_id):
    """Emite um convidado Cortesia/VIP (Modelo A) pelo Admin."""
    event = db.get_event(event_id)
    if not event:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    payload = request.get_json(force=True, silent=True) or {}
    return _issue_courtesy(event_id, payload)


@app.route("/api/organizador/porteiros", methods=["GET"])
@role_required_api("organizador")
def api_organizador_list_porteiros():
    """Lista os porteiros que ESTE Organizador já criou ('porteiro do promotor')."""
    current_user = get_current_user()
    try:
        porteiros = db.list_porteiros_by_organizador(current_user["id"])
        return jsonify({"success": True, "data": porteiros, "limit": config.MAX_PORTEIROS_POR_ORGANIZADOR})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/organizador/porteiros", methods=["POST"])
@role_required_api("organizador")
def api_organizador_create_porteiro():
    """
    Cria um porteiro vinculado a ESTE Organizador ('porteiro do promotor'
    — Fase 5, item C4): só valida check-in dos eventos deste promotor
    (ver `_porteiro_can_access_event`). Limite de 3 por Organizador.
    """
    current_user = get_current_user()
    try:
        existing_count = db.count_porteiros_by_organizador(current_user["id"])
        if existing_count >= config.MAX_PORTEIROS_POR_ORGANIZADOR:
            return jsonify({
                "success": False,
                "error": f"Limite de {config.MAX_PORTEIROS_POR_ORGANIZADOR} porteiros por Organizador atingido.",
            }), 409

        payload = request.get_json(force=True, silent=True) or {}
        full_name = (payload.get("full_name") or "").strip()
        username = (payload.get("username") or "").strip()
        password = payload.get("password") or ""

        if not username:
            return jsonify({"success": False, "error": "O usuário (login) é obrigatório."}), 400
        if len(password) < MIN_PASSWORD_LENGTH:
            return jsonify({
                "success": False,
                "error": f"A senha provisória deve ter pelo menos {MIN_PASSWORD_LENGTH} caracteres.",
            }), 400
        if db.username_exists(username):
            return jsonify({"success": False, "error": "Já existe um usuário com esse login."}), 409

        password_hash = generate_password_hash(password)
        user = db.create_user(
            username, password_hash, "porteiro", full_name=full_name, organizador_id=current_user["id"],
        )
        user.pop("password_hash", None)
        return jsonify({"success": True, "data": user}), 201
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/events/template-xlsx", methods=["GET"])
@role_required_api("admin", "organizador")
def api_download_guest_template():
    """
    Gera e retorna um .xlsx modelo com as colunas EXATAS exigidas pelo
    validador de import (_parse_guests_xlsx / REQUIRED_XLSX_COLUMNS) —
    incluindo Email e Telefone, que são obrigatórios de verdade no
    sistema hoje, mesmo que a Fase 2 tenha pedido um modelo mais
    enxuto (Nome, Cargo/Tipo, Mesa). Gerar o modelo enxuto faria o
    próprio Organizador levar erro de validação na primeira tentativa
    de importar o arquivo que o sistema mandou baixar.
    """
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Convidados"

        headers = config.REQUIRED_XLSX_COLUMNS + [config.OPTIONAL_XLSX_COLUMN_TABLE]
        sheet.append(headers)

        header_fill = PatternFill(start_color="004E92", end_color="004E92", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, len(headers) + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font

        sheet.append(["Ex: João Manuel Sitoe", "joao@email.com", "841234567", "Convidado", "Mesa 01"])

        column_widths = [28, 30, 16, 20, 14]
        for i, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[chr(64 + i)].width = width

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        return send_file(
            buffer,
            as_attachment=True,
            download_name="modelo_convidados_gateflow.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:  # noqa: BLE001
        app.logger.error("Erro ao gerar template xlsx: %s", traceback.format_exc())
        return jsonify({"success": False, "error": f"Erro ao gerar o modelo: {exc}"}), 500


@app.route("/api/events", methods=["POST"])
@role_required_api("admin")
def api_create_event():
    try:
        payload = request.get_json(force=True, silent=True) or {}
        name = (payload.get("name") or "").strip()
        location = (payload.get("location") or "").strip()
        description = (payload.get("description") or "").strip()
        event_date = (payload.get("event_date") or "").strip()

        if not name:
            return jsonify({"success": False, "error": "O nome do evento é obrigatório."}), 400
        if not event_date:
            return jsonify({"success": False, "error": "A data e hora do evento são obrigatórias."}), 400

        event = db.create_event(name, location, description, event_date)
        return jsonify({"success": True, "data": event}), 201
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/events/<event_id>", methods=["GET"])
@role_required_api("admin")
def api_get_event(event_id):
    try:
        event = db.get_event(event_id)
        if not event:
            return jsonify({"success": False, "error": "Evento não encontrado."}), 404
        event["stats"] = db.get_event_stats(event_id)
        event["status"] = event_status.compute_event_status(event.get("event_date"))
        return jsonify({"success": True, "data": event})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/events/<event_id>", methods=["PUT"])
@role_required_api("admin")
def api_update_event(event_id):
    try:
        current_event = db.get_event(event_id)
        if not current_event:
            return jsonify({"success": False, "error": "Evento não encontrado."}), 404
        if event_status.is_past(current_event.get("event_date")):
            return jsonify({
                "success": False,
                "error": "Este evento já foi encerrado. Edição bloqueada, apenas relatórios continuam disponíveis.",
                "event_ended": True,
            }), 403

        payload = request.get_json(force=True, silent=True) or {}
        name = (payload.get("name") or "").strip()
        location = (payload.get("location") or "").strip()
        description = (payload.get("description") or "").strip()
        event_date = (payload.get("event_date") or "").strip()

        if not name:
            return jsonify({"success": False, "error": "O nome do evento é obrigatório."}), 400
        if not event_date:
            return jsonify({"success": False, "error": "A data e hora do evento são obrigatórias."}), 400

        event = db.update_event(event_id, name, location, description, event_date)
        if not event:
            return jsonify({"success": False, "error": "Evento não encontrado."}), 404
        event["status"] = event_status.compute_event_status(event.get("event_date"))
        return jsonify({"success": True, "data": event})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/events/<event_id>", methods=["DELETE"])
@role_required_api("admin")
def api_delete_event(event_id):
    try:
        deleted = db.delete_event(event_id)
        if not deleted:
            return jsonify({"success": False, "error": "Evento não encontrado."}), 404
        return jsonify({"success": True})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


# --------------------------------------------------------------------------
# API: IMPORTAÇÃO DE CONVIDADOS (.xlsx)
# --------------------------------------------------------------------------

@app.route("/api/events/<event_id>/import", methods=["POST"])
@role_required_api("admin")
def api_import_guests(event_id):
    event = db.get_event(event_id)
    if not event:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    guard = _require_module_a_event(event)
    if guard:
        return guard

    if "file" not in request.files:
        return jsonify({"success": False, "error": "Nenhum arquivo enviado."}), 400

    file = request.files["file"]
    if file.filename == "":
        return jsonify({"success": False, "error": "Nome de arquivo inválido."}), 400

    filename = secure_filename(file.filename)
    _, ext = os.path.splitext(filename.lower())
    if ext not in config.ALLOWED_UPLOAD_EXTENSIONS:
        return jsonify({"success": False, "error": "Apenas arquivos .xlsx são aceitos."}), 400

    temp_path = os.path.join(config.UPLOADS_DIR, filename)
    try:
        file.save(temp_path)
        guests_data, warnings = _parse_guests_xlsx(temp_path)

        if not guests_data:
            return jsonify({
                "success": False,
                "error": "Nenhum convidado válido encontrado na planilha."
            }), 400

        # Deduplicação: convidados já existentes para ESTE evento (por
        # e-mail, ou por Nome+Telefone quando não há e-mail) são ignorados
        # em vez de duplicados.
        created_guests, skipped_count = db.import_guests_with_dedup(event_id, guests_data)

        # Gera o QR Code apenas dos convidados efetivamente novos
        generation_errors = []
        for guest in created_guests:
            try:
                qr_path = generate_guest_qrcode(event_id, guest["id"])
                db.update_guest_qr_path(guest["id"], qr_path)
            except QRCodeGenerationError as exc:
                generation_errors.append(str(exc))

        summary = (
            f"{len(created_guests)} convidado(s) adicionado(s), "
            f"{skipped_count} pulado(s) por já existirem neste evento."
        )

        return jsonify({
            "success": True,
            "data": {
                "imported_count": len(created_guests),
                "skipped_count": skipped_count,
                "summary": summary,
                "warnings": warnings,
                "qr_generation_errors": generation_errors,
            },
        }), 201

    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except ValueError as exc:
        return jsonify({"success": False, "error": str(exc)}), 400
    except Exception as exc:  # noqa: BLE001
        app.logger.error("Erro na importação: %s", traceback.format_exc())
        return jsonify({"success": False, "error": f"Erro ao processar a planilha: {exc}"}), 500
    finally:
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass


def _parse_guests_xlsx(file_path):
    """
    Lê a planilha .xlsx e retorna (guests_data, warnings).
    Valida cabeçalho e ignora linhas em branco ou incompletas (registrando
    aviso), sem interromper a importação inteira por causa de uma linha ruim.
    """
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ValueError(f"Arquivo .xlsx corrompido ou inválido: {exc}") from exc

    sheet = workbook.active
    header_row = [str(cell.value).strip() if cell.value is not None else "" for cell in sheet[1]]

    # Mapa auxiliar para localizar colunas de forma insensível a maiúsculas/
    # minúsculas (ex: "Mesa", "MESA" e "mesa" devem ser todos reconhecidos).
    # Em caso de cabeçalhos duplicados (mesmo nome em mais de uma coluna),
    # fica a primeira ocorrência, que é o comportamento mais previsível.
    header_lookup = {}
    for idx, col_name in enumerate(header_row):
        key = col_name.strip().lower()
        if key and key not in header_lookup:
            header_lookup[key] = idx

    def _find_column_index(expected_name):
        return header_lookup.get(expected_name.strip().lower())

    column_index = {}
    for expected_col in config.REQUIRED_XLSX_COLUMNS:
        found_idx = _find_column_index(expected_col)
        if found_idx is None:
            raise ValueError(
                f"Coluna obrigatória ausente na planilha: '{expected_col}'. "
                f"Colunas esperadas: {', '.join(config.REQUIRED_XLSX_COLUMNS)}"
            )
        column_index[expected_col] = found_idx

    # Coluna "Mesa" é OPCIONAL: planilhas antigas (sem essa coluna) continuam
    # funcionando normalmente — todo convidado recebe o valor padrão. A busca
    # também é insensível a maiúsculas/minúsculas (aceita "Mesa", "MESA", "mesa").
    table_col_idx = _find_column_index(config.OPTIONAL_XLSX_COLUMN_TABLE)
    has_table_column = table_col_idx is not None
    if has_table_column:
        column_index[config.OPTIONAL_XLSX_COLUMN_TABLE] = table_col_idx

    guests = []
    warnings = []
    for row_number, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row]
        if all(v is None or str(v).strip() == "" for v in values):
            continue  # linha totalmente vazia, ignora silenciosamente

        try:
            full_name = str(values[column_index["Nome Completo"]] or "").strip()
        except IndexError:
            full_name = ""

        if not full_name:
            warnings.append(f"Linha {row_number}: ignorada por não ter 'Nome Completo'.")
            continue

        def _safe_get(col_name):
            idx = column_index[col_name]
            if idx >= len(values) or values[idx] is None:
                return ""
            return str(values[idx]).strip()

        table_number = _safe_get(config.OPTIONAL_XLSX_COLUMN_TABLE) if has_table_column else ""

        guests.append({
            "full_name": full_name,
            "email": _safe_get("Email"),
            "phone": _safe_get("Telefone"),
            "role": _safe_get("Cargo/Tipo"),
            "table_number": table_number or config.DEFAULT_TABLE_LABEL,
        })

    return guests, warnings


# --------------------------------------------------------------------------
# API: CONVIDADOS
# --------------------------------------------------------------------------

@app.route("/api/events/<event_id>/guests", methods=["POST"])
@role_required_api("admin")
def api_create_guest_manual(event_id):
    """Cadastro manual de UM convidado pelo Admin (fora do fluxo de planilha)."""
    event = db.get_event(event_id)
    if not event:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    guard = _require_module_a_event(event)
    if guard:
        return guard

    try:
        payload = request.get_json(force=True, silent=True) or {}
        full_name = (payload.get("full_name") or "").strip()
        email = (payload.get("email") or "").strip()
        phone = (payload.get("phone") or "").strip()
        role = (payload.get("role") or "").strip()
        table_number = (payload.get("table_number") or "").strip()

        if not full_name:
            return jsonify({"success": False, "error": "O nome completo é obrigatório."}), 400

        guest, is_duplicate = db.create_single_guest(
            event_id, full_name, email, phone, role, table_number
        )

        if is_duplicate:
            return jsonify({
                "success": False,
                "error": f"Já existe um convidado igual neste evento: {guest['full_name']}.",
                "data": guest,
            }), 409

        try:
            qr_path = generate_guest_qrcode(event_id, guest["id"])
            db.update_guest_qr_path(guest["id"], qr_path)
            guest["qr_code_path"] = qr_path
        except QRCodeGenerationError as exc:
            app.logger.error("Falha ao gerar QR Code do convidado manual: %s", exc)

        return jsonify({"success": True, "data": guest}), 201
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except Exception as exc:  # noqa: BLE001 — garante resposta JSON mesmo em erro inesperado
        app.logger.error("Erro inesperado no cadastro manual de convidado: %s", traceback.format_exc())
        return jsonify({"success": False, "error": f"Erro ao cadastrar convidado: {exc}"}), 500


@app.route("/api/events/<event_id>/guests/<guest_id>", methods=["PUT"])
@role_required_api("admin")
def api_update_guest(event_id, guest_id):
    event = db.get_event(event_id)
    if not event:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    guard = _require_module_a_event(event)
    if guard:
        return guard

    try:
        payload = request.get_json(force=True, silent=True) or {}
        full_name = (payload.get("full_name") or "").strip()
        email = (payload.get("email") or "").strip()
        phone = (payload.get("phone") or "").strip()
        role = (payload.get("role") or "").strip()
        table_number = (payload.get("table_number") or "").strip()

        if not full_name:
            return jsonify({"success": False, "error": "O nome completo é obrigatório."}), 400

        guest, is_duplicate = db.update_guest(
            event_id, guest_id, full_name, email, phone, role, table_number
        )

        if guest is None and not is_duplicate:
            return jsonify({"success": False, "error": "Convidado não encontrado."}), 404
        if is_duplicate:
            return jsonify({
                "success": False,
                "error": f"Já existe outro convidado igual neste evento: {guest['full_name']}.",
                "data": guest,
            }), 409

        return jsonify({"success": True, "data": guest})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except Exception as exc:  # noqa: BLE001
        app.logger.error("Erro inesperado ao editar convidado: %s", traceback.format_exc())
        return jsonify({"success": False, "error": f"Erro ao editar convidado: {exc}"}), 500


@app.route("/api/events/<event_id>/guests/<guest_id>", methods=["DELETE"])
@role_required_api("admin")
def api_delete_guest(event_id, guest_id):
    event = db.get_event(event_id)
    if not event:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    guard = _require_module_a_event(event)
    if guard:
        return guard

    try:
        guest = db.get_guest_by_event(event_id, guest_id)
        deleted = db.delete_guest(event_id, guest_id)
        if not deleted:
            return jsonify({"success": False, "error": "Convidado não encontrado."}), 404

        # Remove o arquivo de QR Code órfão do disco, se existir (não crítico se falhar)
        if guest and guest.get("qr_code_path") and os.path.exists(guest["qr_code_path"]):
            try:
                os.remove(guest["qr_code_path"])
            except OSError:
                pass

        return jsonify({"success": True})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/events/<event_id>/guests", methods=["GET"])
@role_required_api("admin", "porteiro")
def api_list_guests(event_id):
    event = db.get_event(event_id)
    if not event:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    current_user = get_current_user()
    if current_user["role"] == "porteiro" and not _porteiro_can_access_event(current_user, event):
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    try:
        search = request.args.get("search", "").strip()
        guests = db.list_guests(event_id, search=search or None)
        stats = db.get_event_stats(event_id)
        return jsonify({"success": True, "data": guests, "stats": stats})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/events/<event_id>/guests/<guest_id>/qrcode", methods=["GET"])
@role_required_api("admin", "porteiro")
def api_get_guest_qrcode(event_id, guest_id):
    current_user = get_current_user()
    if current_user["role"] == "porteiro":
        event = db.get_event(event_id)
        if not _porteiro_can_access_event(current_user, event):
            abort(404)
    guest = db.get_guest_by_event(event_id, guest_id)
    if not guest:
        abort(404)
    # Não confiamos apenas no caminho salvo em `qr_code_path`: em produção
    # serverless, o arquivo pode ter sido gerado num cold start anterior
    # cujo /tmp já não existe mais. Regeneramos sob demanda quando o
    # arquivo não está (mais) no disco -- o conteúdo do QR Code é sempre
    # o mesmo (deriva só do guest_id), então isso é 100% seguro.
    qr_path = guest.get("qr_code_path")
    if not qr_path or not os.path.exists(qr_path):
        try:
            qr_path = generate_guest_qrcode(event_id, guest_id)
            db.update_guest_qr_path(guest_id, qr_path)
        except QRCodeGenerationError:
            abort(404)
    return send_file(qr_path, mimetype="image/png")


@app.route("/api/events/<event_id>/guests/export-pdf", methods=["GET"])
@role_required_api("admin")
def api_export_pdf(event_id):
    event = db.get_event(event_id)
    if not event:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    try:
        guests = db.list_guests(event_id)
        if not guests:
            return jsonify({"success": False, "error": "Este evento não possui convidados."}), 400

        pdf_path = generate_invites_pdf(event, guests)

        # Nome de arquivo dinâmico baseado no nome do EVENTO selecionado,
        # limpo de acentos/espaços/caracteres especiais para evitar erros
        # no sistema de arquivos do navegador/SO ao salvar o download.
        # Ex: "Workshop de Inovação 2026" -> "Workshop_de_Inovacao_2026.pdf"
        event_slug = slugify_filename(event["name"], fallback=event_id)
        download_name = f"{event_slug}.pdf"

        response = send_file(
            pdf_path,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/pdf",
        )
        # Reforça explicitamente o Content-Disposition com o nome dinâmico
        # (send_file já faz isso via download_name, mas deixamos explícito
        # aqui para garantir compatibilidade entre navegadores/proxies).
        response.headers["Content-Disposition"] = f'attachment; filename="{download_name}"'
        return response
    except PDFGenerationError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/events/<event_id>/guests/contingency-pdf", methods=["GET"])
@role_required_api("admin")
def api_export_contingency_pdf(event_id):
    """
    PDF de contingência: lista minimalista em ordem alfabética (Nome +
    Mesa), sem QR Code, para impressão de emergência caso o sistema
    fique indisponível durante o evento.
    """
    event = db.get_event(event_id)
    if not event:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    try:
        guests = db.list_guests_alphabetical(event_id)
        if not guests:
            return jsonify({"success": False, "error": "Este evento não possui convidados."}), 400

        pdf_path = generate_contingency_pdf(event, guests)

        event_slug = slugify_filename(event["name"], fallback=event_id)
        download_name = f"{event_slug}_contingencia.pdf"

        response = send_file(
            pdf_path,
            as_attachment=True,
            download_name=download_name,
            mimetype="application/pdf",
        )
        response.headers["Content-Disposition"] = f'attachment; filename="{download_name}"'
        return response
    except PDFGenerationError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/events/<event_id>/guests/attendance-report", methods=["GET"])
@role_required_api("admin")
def api_export_attendance_report(event_id):
    """Relatório de presença em CSV (quem compareceu / quem faltou) — útil após o evento encerrar."""
    event = db.get_event(event_id)
    if not event:
        return jsonify({"success": False, "error": "Evento não encontrado."}), 404
    try:
        guests = db.list_guests(event_id)
        if not guests:
            return jsonify({"success": False, "error": "Este evento não possui convidados."}), 400

        csv_content = generate_attendance_report_csv(event, guests)
        event_slug = slugify_filename(event["name"], fallback=event_id)
        download_name = f"{event_slug}_relatorio_presenca.csv"

        # BOM (\ufeff) garante que acentos abram corretamente no Excel
        csv_bytes = ("\ufeff" + csv_content).encode("utf-8")

        response = app.response_class(csv_bytes, mimetype="text/csv")
        response.headers["Content-Disposition"] = f'attachment; filename="{download_name}"'
        return response
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


# --------------------------------------------------------------------------
# API: CHECK-IN
# --------------------------------------------------------------------------

@app.route("/api/checkin/manual/<guest_id>", methods=["POST"])
@role_required_api("admin", "porteiro")
def api_checkin_manual(guest_id):
    try:
        guest_preview = db.get_guest(guest_id)
        if not guest_preview:
            return jsonify({"success": False, "error": "Convidado não encontrado."}), 404

        event = db.get_event(guest_preview["event_id"])
        if not _porteiro_can_access_event(get_current_user(), event):
            return jsonify({"success": False, "error": "Convidado não encontrado."}), 404
        if event and event_status.is_past(event.get("event_date")):
            return jsonify({
                "success": False,
                "error": "Este evento já foi encerrado. Não é possível fazer novos check-ins.",
                "event_ended": True,
            }), 403

        # O responsável pelo check-in vem da sessão autenticada, não mais
        # de um campo livre enviado pelo cliente (evita forjar o nome).
        checkin_by = session.get("username", "desconhecido")

        success, message, guest = db.checkin_guest(guest_id, checkin_by=checkin_by)
        if message == "not_found":
            return jsonify({"success": False, "error": "Convidado não encontrado."}), 404
        if message == "already_checked_in":
            return jsonify({
                "success": False,
                "error": "Check-in já realizado anteriormente!",
                "data": guest,
                "already_checked_in": True,
            }), 409
        return jsonify({"success": True, "data": guest})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/checkin/qr", methods=["POST"])
@role_required_api("admin", "porteiro")
def api_checkin_qr():
    try:
        payload = request.get_json(force=True, silent=True) or {}
        raw_qr_value = payload.get("qr_value", "")
        event_id = payload.get("event_id")
        # O responsável pelo check-in vem da sessão autenticada
        checkin_by = session.get("username", "desconhecido")

        if not event_id:
            return jsonify({"success": False, "error": "event_id é obrigatório."}), 400

        event = db.get_event(event_id)
        if not _porteiro_can_access_event(get_current_user(), event):
            return jsonify({"success": False, "error": "Evento não encontrado."}), 404
        if event and event_status.is_past(event.get("event_date")):
            return jsonify({
                "success": False,
                "error": "Este evento já foi encerrado. Não é possível fazer novos check-ins.",
                "event_ended": True,
            }), 403

        guest_id = extract_guest_id_from_payload(raw_qr_value)
        if not guest_id:
            return jsonify({
                "success": False,
                "error": "QR Code inválido ou não pertence a este sistema."
            }), 400

        guest = db.get_guest_by_event(event_id, guest_id)
        if not guest:
            return jsonify({
                "success": False,
                "error": "QR Code não corresponde a nenhum convidado deste evento."
            }), 404

        success, message, updated_guest = db.checkin_guest(guest_id, checkin_by=checkin_by)
        if message == "already_checked_in":
            return jsonify({
                "success": False,
                "error": f"Check-in já realizado! ({updated_guest['full_name']})",
                "data": updated_guest,
                "already_checked_in": True,
            }), 409

        return jsonify({"success": True, "data": updated_guest})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


# --------------------------------------------------------------------------
# API: GERENCIAMENTO DE USUÁRIOS (somente admin)
# --------------------------------------------------------------------------

VALID_ROLES = ("admin", "porteiro")
MIN_PASSWORD_LENGTH = 8


@app.route("/api/admin/users", methods=["GET"])
@role_required_api("admin")
def api_list_users():
    try:
        users = db.list_users()
        return jsonify({"success": True, "data": users})
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/admin/users", methods=["POST"])
@role_required_api("admin")
def api_create_user():
    """
    Cria um novo usuário (admin ou porteiro). Qualquer usuário autenticado
    com papel 'admin' pode criar novos perfis — não há distinção especial
    de "super admin": o primeiro admin (criado automaticamente no primeiro
    boot) tem exatamente as mesmas permissões que qualquer outro admin
    criado depois.
    """
    try:
        payload = request.get_json(force=True, silent=True) or {}
        full_name = (payload.get("full_name") or "").strip()
        username = (payload.get("username") or "").strip()
        password = payload.get("password") or ""
        role = (payload.get("role") or "").strip().lower()

        if not username:
            return jsonify({"success": False, "error": "O usuário (login) é obrigatório."}), 400
        if role not in VALID_ROLES:
            return jsonify({"success": False, "error": "Tipo inválido. Use 'admin' ou 'porteiro'."}), 400
        if len(password) < MIN_PASSWORD_LENGTH:
            return jsonify({
                "success": False,
                "error": f"A senha provisória deve ter pelo menos {MIN_PASSWORD_LENGTH} caracteres.",
            }), 400
        if db.username_exists(username):
            return jsonify({"success": False, "error": "Já existe um usuário com esse login."}), 409

        password_hash = generate_password_hash(password)
        user = db.create_user(username, password_hash, role, full_name=full_name)
        user.pop("password_hash", None)  # nunca retorna o hash ao cliente
        return jsonify({"success": True, "data": user}), 201
    except db.DatabaseError as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route("/api/admin/security/invalidate-all-sessions", methods=["POST"])
@role_required_api("admin")
def api_invalidate_all_sessions():
    """
    "Kill switch" de sessões: gera uma nova chave secreta e a torna ativa
    imediatamente, invalidando TODAS as sessões existentes (inclusive a
    do próprio admin que chamou esta rota — ele também precisará logar de
    novo). Útil para encerrar o acesso de todos os aparelhos de uma vez
    (ex: fim de um evento, dispositivo perdido/roubado, ou depois de
    testes, antes de entregar os tablets para uso real).
    """
    try:
        new_key = os.urandom(32).hex()
        with open(config.SECRET_KEY_PATH, "w", encoding="utf-8") as f:
            f.write(new_key)
        app.secret_key = new_key

        # IMPORTANTE: como a sessão é "permanente" (renovada a cada
        # requisição), o Flask re-assinaria automaticamente o cookie
        # desta própria resposta com a chave NOVA antes de enviá-la —
        # o que manteria o admin que chamou esta rota ainda logado.
        # Limpamos a sessão explicitamente para fechar essa brecha
        # também para quem está fazendo a chamada agora.
        session.clear()

        return jsonify({
            "success": True,
            "message": "Todas as sessões foram encerradas. Será necessário fazer login novamente em todos os aparelhos.",
        })
    except OSError as exc:
        return jsonify({"success": False, "error": f"Erro ao rotacionar a chave de sessão: {exc}"}), 500


if __name__ == "__main__":
    # HTTPS automático é OPCIONAL (desligado por padrão). Em alguns Chrome
    # — principalmente em computadores/telemóveis com política corporativa
    # ou gerenciados por uma empresa/escola — o botão "Continuar mesmo
    # assim" do aviso de certificado autoassinado fica bloqueado e a
    # página nunca carrega. Para esses casos, o caminho mais confiável é
    # liberar a câmera diretamente no navegador (ver README, seção de
    # câmera), sem depender de certificado nenhum.
    #
    # Para ligar o HTTPS automático (se quiser tentar mesmo assim), rode:
    #   CHECKIN_ENABLE_HTTPS=1 python3 app.py
    enable_https = os.environ.get("CHECKIN_ENABLE_HTTPS", "").strip() in ("1", "true", "True")

    cert_path, key_path = (None, None)
    if enable_https:
        cert_path, key_path = ensure_self_signed_cert()

    if cert_path and key_path:
        scheme = "https"
        ssl_context = (cert_path, key_path)
    else:
        scheme = "http"
        ssl_context = None

    print(f"\n>> Sistema de Check-in disponível na rede local em: {scheme}://<IP-DESTE-COMPUTADOR>:{config.PORT}")

    if scheme == "https":
        print(">> IMPORTANTE: o navegador vai mostrar um aviso de 'conexão não segura' na")
        print(">> primeira vez que cada tablet acessar — isso é ESPERADO (é um certificado")
        print(">> autoassinado, gerado só para permitir o uso da câmera). Toque em 'Avançado'")
        print(">> e depois em 'Acessar mesmo assim' / 'Prosseguir'. Isso só aparece uma vez.")
    else:
        print(">> Rodando em HTTP. Para a câmera do leitor de QR Code funcionar em cada")
        print(">> tablet/computador, configure UMA VEZ por aparelho (veja o README,")
        print(">> seção 'Câmera não funciona pelo IP da rede'):")
        print(">>   chrome://flags/#unsafely-treat-insecure-origin-as-secure")

    print(">> Painel Administrativo (Multi-Eventos): /admin   |   Painel do Porteiro (Multi-Eventos): /checkin\n")
    app.run(host=config.HOST, port=config.PORT, debug=False, threaded=True, ssl_context=ssl_context)
